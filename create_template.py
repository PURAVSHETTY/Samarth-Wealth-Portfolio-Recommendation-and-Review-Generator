"""
create_template.py
Generates a sample Excel input template for the portfolio proposal generator.
Run this once to get a blank template you can fill in.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def hex_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


NAVY = "0D1B4B"
GOLD = "C9A84C"
WHITE = "FFFFFF"
LIGHT_GREY = "F5F5F5"


def style_header(cell, bg=NAVY, fg=WHITE, bold=True, size=11):
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill = hex_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()


def style_data(cell, bold=False, bg=WHITE, fg="000000", align="left"):
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=10)
    cell.fill = hex_fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()


wb = Workbook()

# ── Sheet 1: Firm Info ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Firm_Info"
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 40

headers = [
    ("FIELD", "VALUE"),
    ("Firm Name", "Samarth Wealth Pvt. Ltd."),
    ("Tagline", "Clients First. Always and Everytime"),
    ("Founded", "2011"),
    ("Team Size", "30"),
    ("AUM (Crores)", "1500"),
    ("Location", "Thane, Maharashtra"),
    ("Address", "G-75/76, Eternity Commercial Premises, Teen Haat Naka, Off LBS Marg, Thane (W) - 400604"),
    ("Email", "relationships@samarthedufin.com"),
    ("Website", "www.samarthwealth.in"),
    ("Phone", "+91 7738245239 | 7208571404"),
    ("AMFI ARN", "ARN-286847"),
    ("Advisor Name", "Abhinandan Honale"),
    ("Advisor Title", "Co-Founder & Head: Wealth Management"),
    ("Advisor Credentials", "FRM, MBA Finance, B.E."),
    ("Advisor Background", "Ex-ICICI Bank Wealth Management, Deutsche Bank Investment Banking"),
]

for i, (field, value) in enumerate(headers, 1):
    ws1[f"A{i}"] = field
    ws1[f"B{i}"] = value
    if i == 1:
        style_header(ws1[f"A{i}"])
        style_header(ws1[f"B{i}"])
    else:
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        style_data(ws1[f"A{i}"], bold=True, bg=bg)
        style_data(ws1[f"B{i}"], bg=bg)
ws1.row_dimensions[1].height = 22

# ── Sheet 2: Client Info ────────────────────────────────────────────────────
ws2 = wb.create_sheet("Client_Info")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 40

client_fields = [
    ("FIELD", "VALUE"),
    ("Client Name", "Mrs. Anu Shivdasani & Family"),
    ("Portfolio Corpus (INR)", "12,00,00,000"),
    ("Report Date", "May 2026"),
    ("Investment Horizon", "5+ Years"),
    ("Risk Profile", "Moderate-Aggressive"),
    ("Primary Objective", "Wealth Compounding with Regular Income"),
    ("Tax Bracket", "30%"),
    ("Executive Briefing", "This is a sample custom executive briefing from Excel. It will be displayed directly on Page 4 without generating any AI narratives."),
    ("Portfolio Thesis & Market Overview", "This is a sample custom portfolio thesis and market overview from Excel. It will be displayed directly on Page 4 without generating any AI narratives.")
]

for i, (field, value) in enumerate(client_fields, 1):
    ws2[f"A{i}"] = field
    ws2[f"B{i}"] = value
    if i == 1:
        style_header(ws2[f"A{i}"])
        style_header(ws2[f"B{i}"])
    else:
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        style_data(ws2[f"A{i}"], bold=True, bg=bg)
        style_data(ws2[f"B{i}"], bg=bg)

# ── Sheet 3: Asset Allocation ───────────────────────────────────────────────
ws3 = wb.create_sheet("Asset_Allocation")
ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 50

alloc_headers = ["Part", "Segment Name", "Allocation %", "Objective"]
for col, h in enumerate(alloc_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    style_header(cell)

alloc_data = [
    (1, "Safety Reserve", 10, "Keeps a portion of your money safe, highly stable, and easily accessible whenever needed."),
    (2, "Stable Income", 25, "Provides steady, regular payouts with lower volatility to protect your capital."),
    (3, "Balanced Growth", 25, "Helps grow wealth while reducing risk during market ups and downs."),
    (4, "Wealth Creation", 40, "Aims for high long-term growth to build wealth over the coming years."),
]

for i, row in enumerate(alloc_data, 2):
    for col, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        style_data(cell, bg=LIGHT_GREY if i % 2 == 0 else WHITE,
                   align="center" if col in (1, 3) else "left")
ws3.row_dimensions[1].height = 22

# ── Sheet 4: Products ───────────────────────────────────────────────────────
ws4 = wb.create_sheet("Products")
for col, width in enumerate([5, 12, 32, 22, 18, 18, 60], 1):
    ws4.column_dimensions[get_column_letter(col)].width = width

prod_headers = ["Part", "Segment", "Product Name", "Asset Class", "Allocation (INR)", "Target Return", "Core Rationale"]
for col, h in enumerate(prod_headers, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    style_header(cell)

products = [
    # Part, Segment, Product, Asset Class, Allocation, Return, Rationale
    (1, "Safety Reserve", "SBI Liquid Fund", "Mutual Fund", "60,00,000", "6.5% - 7.5%",
     "Keeps your cash safe, highly stable, and ready for immediate withdrawal when needed."),
    (1, "Safety Reserve", "ICICI Prudential Arbitrage Fund", "Mutual Fund", "60,00,000", "7% - 8%",
     "Capitalizes on short-term price differences in shares to yield tax-friendly, low-risk returns."),
    (2, "Stable Income", "ICICI Prudential Corporate Bond Fund", "Mutual Fund", "1,50,00,000", "7.5% - 8.5%",
     "Holds debt from high-quality companies to secure reliable interest payouts with minimal risk."),
    (2, "Stable Income", "Kotak Debt Hybrid Fund", "Mutual Fund", "1,50,00,000", "8% - 9%",
     "Combines bond stability with a small equity slice to shield cash from inflation while paying regular yields."),
    (3, "Balanced Growth", "SBI Equity Hybrid Fund", "Mutual Fund", "1,50,00,000", "11% - 13%",
     "Maintains a balanced mix of equities and debt to grow capital while cushioning against market dips."),
    (3, "Balanced Growth", "HDFC Multi-Asset Allocation Fund", "Mutual Fund", "1,50,00,000", "12% - 14%",
     "Spreads wealth across gold, bonds, and shares to ensure a smooth and steady growth path."),
    (4, "Wealth Creation", "Parag Parikh Flexi Cap Fund", "Mutual Fund", "1,80,00,000", "14% - 16%",
     "Invests in a select group of leading Indian and international businesses for reliable wealth expansion."),
    (4, "Wealth Creation", "Nippon India Small Cap Fund", "Mutual Fund", "1,50,00,000", "16% - 18%",
     "Backs emerging, fast-growing local companies to capture strong compounding gains over the long term."),
    (4, "Wealth Creation", "Mirae Asset Large Cap Fund", "Mutual Fund", "1,50,00,000", "12% - 14%",
     "Holds dominant industry leaders to provide a stable, solid base for long-term growth.")
]

for i, row in enumerate(products, 2):
    for col, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=col, value=val)
        style_data(cell, bg=LIGHT_GREY if i % 2 == 0 else WHITE,
                   align="center" if col in (1, 5, 6) else "left")
ws4.row_dimensions[1].height = 22
for r in range(2, len(products) + 2):
    ws4.row_dimensions[r].height = 50

wb.save("CLIENT_TEMPLATE.xlsx")
print("Template saved: CLIENT_TEMPLATE.xlsx")
